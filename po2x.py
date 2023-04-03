import openpyxl
import polib

def po2x(po_file, xlsx_file):
    po = polib.pofile(po_file)
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = po_file.stem

    ws.cell(row=1, column=1, value="reference_comment")
    ws.cell(row=1, column=2, value="msgctxt")
    ws.cell(row=1, column=3, value="msgid")
    ws.cell(row=1, column=4, value="msgstr")

    for i, entry in enumerate(po, start=2):
        ws.cell(row=i, column=1, value=entry.occurrences[0][0] if entry.occurrences else '').number_format = '@'
        ws.cell(row=i, column=2, value=entry.msgctxt).number_format = '@'
        
        source = entry.msgid
        translation = entry.msgstr

        if len(source) > 32767:
            source = source[:32766] + " !!ERR CONVERTED!!"
        if len(translation) > 32767:
            translation = translation[:32766] + " !!ERR CONVERTED!!"

        ws.cell(row=i, column=3, value=source).number_format = '@'
        ws.cell(row=i, column=4, value=translation).number_format = '@'

        # Debug print statement for problematic entries
        if len(entry.msgid) != len(source) or len(entry.msgstr) != len(translation):
            print(f"Truncation detected at row {i}:\nOriginal msgid: {entry.msgid}\nConverted msgid: {source}\nOriginal msgstr: {entry.msgstr}\nConverted msgstr: {translation}\n")

    wb.save(xlsx_file)
